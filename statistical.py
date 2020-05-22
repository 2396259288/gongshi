
import pandas as pd
import os
from datetime import datetime
from datetime import date
from datetime import timedelta
from copy import deepcopy
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import shutil
import numpy as np
import traceback
import logging



def getYesterday(): 
    today=date.today() 
    oneday=timedelta(days=1) 
    yesterday=today-oneday  
    return ''.join(str(yesterday).split('-'))
# time_ymd='20200513'
# yesterday_ymd = datetime.strptime(str(getYesterday()), '%Y%m%d')
yesterday_ymd = getYesterday()
time_ymd=datetime.strftime(datetime.now(), '%Y%m%d')
time_ymdHMS = datetime.strftime(datetime.now(), '%Y%m%d %H:%M:%S')
time_hm = datetime.strftime(datetime.now(), '%H:%M')
logging.basicConfig(filename='error.txt', level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')


class WorkRecode():
     
    # def is_datetime(self, dt):
    #     if str(dt) == time_ymd:
    #         return 1
    #     else:
    #         return 0

    def merge(self, config_wb, config_ws, config_path, distribute_path, tem_path, merge_path, error_path):
        logs_ws = config_wb['logs']

        if not os.path.exists(merge_path):
            shutil.copy(tem_path,merge_path)

        #读取文件夹中的xlsx文件
        dflist = []
        print('正在收集文件')
        for file in os.listdir(distribute_path):
            try:
                if file == 'workbook.xlsx':
                    continue
                filename = os.path.join(distribute_path, file)
                df = pd.read_excel(filename, sheet_name='Sheet1', skiprows=([0,1]))
                dflist.append(df)
            except Exception as e:
                print('文件正在使用....添加失败！！', file)
                logs_ws.append(['文件正在使用....添加失败！！'+file, time_ymdHMS])  
                errorFile = open(error_path, 'a')    
                errorFile.write(traceback.format_exc())
                errorFile.close()
                continue

        #存成读取总记录文件
        sum_wb = load_workbook(merge_path)
        sum_ws = sum_wb['Sheet1']
        # sum_ws_logs = sum_wb['logs']
        # sum_df = pd.read_excel('sum.xlsx')
        #读取sum_ws的总行 总列
        rows = sum_ws.max_row
        cols = sum_ws.max_column
        # print(rows)
        time_col_data = []
        for i in range(4, rows+1):
            cell_val = sum_ws.cell(row = i, column = 2).value
            time_col_data.append(str(cell_val))
        if time_ymd in time_col_data:
            for i in range(4, rows+1):
                # print(sum_ws.cell(row = i, column = 4).value)
                if str(sum_ws.cell(row = i, column = 2).value) == time_ymd:
                    sum_ws.delete_rows(i, i+len(time_col_data))
            sum_wb.save(merge_path)
            sum_wb.close()

        sum_wb = load_workbook(merge_path)
        sum_ws = sum_wb['Sheet1']

        flag = 0
        for df in dflist:
            flag = flag+1
            try:
            #文件没填写不插入
                if df.shape[0] == 0:
                    # print(flag)
                    logs_ws.append(['用户%s未填写' % os.listdir(distribute_path)[flag-1].split('.')[0].split('-')[0], time_ymdHMS])       
                    continue
                
                flag_insert = 0
                for i in range(df.shape[0]):
                    if str((df.iloc[i])[1]) == time_ymd:
                        sum_ws.append(list(df.iloc[i]))
                        print('插入成功', i)
                        flag_insert = flag_insert + 1    
                if flag_insert == 0:
                    logs_ws.append(['用户%s未填写' % list(df['姓名'])[0], time_ymdHMS])   
            except Exception as e:
                print(e)
                errorFile = open(error_path, 'a')
                errorFile.write(traceback.format_exc())
                errorFile.close()
                continue
        # logs_ws.auto_filter.ref = "A1:B"+str(logs_ws.max_row)
        # logs_ws.auto_filter.add_sort_condition("B2:"+str(logs_ws.max_row)) 
        config_wb.save(config_path)
        sum_wb.save(merge_path)
        sum_wb.close()
        config_ws.cell(2, 2).value = time_ymd
        print('收集完成')
    
    def distribute(self, config_wb, config_ws, config_path, distribute_path, tem_path, error_path):
        #判断这个字符汉字
        def is_chinese(chr):
            if chr >= '\u4e00' and chr <= '\u9fa5':
                return 0
            else:
                return 1

        #获取当前日期
        now_time = datetime.strftime(datetime.now(), '%Y%m%d')
        #读取名单
        name_df = pd.read_excel(config_path, sheet_name='names')
        #按名单创建 姓名-日期 格式的xlsx文件

        #创建文件夹存放拆分好的excel表
        # if not os.path.exists(distribute_path):
        #     os.mkdir(distribute_path)
        # else:
        #     shutil.rmtree(distribute_path)
        #     os.mkdir(distribute_path)


        for name in name_df['姓名']:
        # 判断 名字不为空 且都是汉字
            try:    
                if name != ' ' and sum([is_chinese(i) for i in str(name)]) == 0:
                    excel_name = str(name)+'-'+now_time
                    shutil.copy(tem_path,distribute_path+'/'+excel_name+'.xlsx')
            except Exception as e:
                print(e)
                errorFile = open(error_path, 'a')
                errorFile.write(traceback.format_exc())
                errorFile.close()
                continue
        config_ws.cell(2, 5).value = 1
        print('分发完成')

    def sum_eachday(self, worklogssum_path, worklogssum_tem_path, sum_eachday_path, worklogssum_config_wb, worklogssum_config_path):
        logs_ws = worklogssum_config_wb['logs']
        sum_eachday_xlsx = os.path.join(sum_eachday_path, time_ymd+'.xlsx')
        if not os.path.exists(sum_eachday_xlsx):
            shutil.copy(worklogssum_tem_path,sum_eachday_xlsx)
        else:
            os.remove(sum_eachday_xlsx)
            shutil.copy(worklogssum_tem_path,sum_eachday_xlsx)


        sum_eachday_wb = load_workbook(sum_eachday_xlsx)
        sum_eachday_ws = sum_eachday_wb['Sheet1']
        for file in os.listdir(worklogssum_path):
            try:
                if file in ['workbook.xlsx', 'config.xlsx']:
                    continue
                filename = os.path.join(worklogssum_path, file)
                df = pd.read_excel(filename, sheet_name='Sheet1', skiprows=([0,1]))
                for i in range(df.shape[0]):
                    if str((df.iloc[i])[1]) == time_ymd:
                            sum_eachday_ws.append(list(df.iloc[i]))
            except Exception as e:
                print('文件正在使用....添加失败！！', file)
                logs_ws.append(['文件正在使用....添加失败！！'+file, time_ymdHMS])  
                errorFile = open(error_path, 'a')    
                errorFile.write(traceback.format_exc())
                errorFile.close()
                continue
        worklogssum_config_ws.cell(2, 3).value = 1
        worklogssum_config_wb.save(worklogssum_config_path)
        worklogssum_config_wb.close()
        sum_eachday_wb.save(sum_eachday_xlsx)
        sum_eachday_wb.close()



if __name__ == '__main__':
    
    current_dir = os.path.dirname(os.path.abspath(__file__))
    current_dir_dir = os.path.dirname(current_dir)
    bumen_list = os.listdir(os.path.join(current_dir_dir, 'worklogs'))
    # distribute_path = os.path.join(current_dir_dir, 'worklogs/'+bumen)
    # config_path = os.path.join(current_dir, 'config.xlsx')
    work_recode = WorkRecode()
    # config_wb = load_workbook(config_path)
    # config_ws = config_wb['config']
    # tem_path = os.path.join(current_dir, config_ws.cell(2, 4).value)
    # merge_path = os.path.join(current_dir_dir, 'worklogssum/'+bumen+'_sum.xlsx')
    error_path = os.path.join(current_dir, 'error.txt')
    worklogssum_path = os.path.join(current_dir_dir, 'worklogssum')
    worklogssum_config_path = os.path.join(worklogssum_path, 'config.xlsx')
    worklogssum_tem_path = os.path.join(worklogssum_path, 'workbook.xlsx')
    worklogssum_config_wb = load_workbook(worklogssum_config_path)
    sum_eachday_path = os.path.join(current_dir_dir, 'sum_eachday')
    worklogssum_config_ws = worklogssum_config_wb['config']
    
    for bumen in bumen_list:
        if os.path.exists(os.path.join(current_dir_dir, 'worklogs/'+bumen+'/config.xlsx')):
            config_path = os.path.join(current_dir_dir, 'worklogs/'+bumen+'/config.xlsx')
            merge_path = os.path.join(current_dir_dir, 'worklogssum/'+bumen+'_sum.xlsx')
            distribute_path = os.path.join(current_dir_dir, 'worklogs/'+bumen)
            config_wb = load_workbook(config_path)
            config_ws = config_wb['config']
            tem_path = os.path.join(current_dir_dir, 'worklogs/'+bumen+'/'+config_ws.cell(2, 4).value)
            if config_ws.cell(2, 5).value == 0:
                if time_hm > str(config_ws.cell(2, 3).value):
                    work_recode.distribute(config_wb, config_ws, config_path, distribute_path, tem_path, error_path)
                else:
                    print('时间未到')
            else:
                print(bumen, '已经完成分发')

            if str(config_ws.cell(2, 2).value) != time_ymd:
                if time_hm > str(config_ws.cell(2, 1).value):
                    work_recode.merge(config_wb, config_ws, config_path, distribute_path, tem_path, merge_path, error_path)
                else:
                    print('时间未到')
            else:
                print(bumen, '今日已经收集')
            config_wb.save(config_path)
            config_wb.close()
        else:
            print('该部门没有上传名单')
    worklogssum_config_ws.cell(2, 2).value =time_ymd
    worklogssum_config_wb.save(worklogssum_config_path)
    worklogssum_config_wb.close()

    #如果config中的 收集标记 列 的值为 0  立即执行 否则次日 8：00 以后执行
    if worklogssum_config_ws.cell(2, 3).value == 0:
        work_recode.sum_eachday(worklogssum_path,worklogssum_tem_path, sum_eachday_path,worklogssum_config_wb, worklogssum_config_path)
    
    elif str(worklogssum_config_ws.cell(2, 2).value) == yesterday_ymd and time_hm > worklogssum_config_ws.cell(2, 1).value:
        work_recode.sum_eachday(worklogssum_path,worklogssum_tem_path, sum_eachday_path,worklogssum_config_wb, worklogssum_config_path)
    else:
        print('时间没到')
